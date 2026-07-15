#!/usr/bin/env python3
"""Build and validate the public restaurant JSON from one Excel or CSV source."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "bangkok_food_combined_ready.xlsx"
DEFAULT_OUTPUT = ROOT / "restaurants.json"
DEFAULT_REPORT = ROOT / "data" / "quality-report.json"

REQUIRED_COLUMNS = {
    "name",
    "lat",
    "lon",
    "rating",
    "review_count",
    "primary_cuisine",
    "price_level",
    "url",
}
OPTIONAL_COLUMNS = {"cuisine_subtype", "address", "image_url", "city", "country"}
EXPORT_COLUMNS = [
    "name",
    "lat",
    "lon",
    "primary_cuisine",
    "cuisine_subtype",
    "rating",
    "weighted_rating",
    "rating_norm",
    "review_count",
    "review_weight_norm",
    "price_min",
    "price_max",
    "price_mid",
    "address",
    "url",
    "image_url",
]


def normalize_text(value: Any, *, collapse_space: bool = True) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split()) if collapse_space else text


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def parse_price(value: Any) -> tuple[float, float, float] | None:
    """Parse values such as 200-400, 200–400, ฿200, or 1,000+."""
    text = normalize_text(value)
    if text is None:
        return None
    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text.replace(",", ""))]
    if not numbers:
        return None
    low = numbers[0]
    high = numbers[1] if len(numbers) > 1 else numbers[0]
    if low > high:
        low, high = high, low
    return low, high, (low + high) / 2


def load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - exercised in deployment setup
            raise RuntimeError("Excel input requires openpyxl; run pip install -r requirements.txt") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = [normalize_text(value) for value in next(iterator)]
        rows = [dict(zip(headers, values)) for values in iterator if any(value is not None for value in values)]
        workbook.close()
        return rows
    raise ValueError(f"Unsupported source type: {suffix}. Use .xlsx, .xlsm, or .csv.")


def source_modified_date(path: Path) -> str | None:
    """Return a stable source date from Excel metadata when available."""
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised in deployment setup
        raise RuntimeError("Excel input requires openpyxl; run pip install -r requirements.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    modified = workbook.properties.modified
    workbook.close()
    return modified.date().isoformat() if modified else None


def clean_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned: list[dict[str, Any]] = []
    for row in rows:
        item = {str(key).strip(): value for key, value in row.items() if key is not None}
        for field in ("name", "primary_cuisine", "cuisine_subtype", "address", "city", "country"):
            item[field] = normalize_text(item.get(field))
        for field in ("url", "image_url"):
            item[field] = normalize_text(item.get(field), collapse_space=False)
        item["lat"] = to_float(item.get("lat"))
        item["lon"] = to_float(item.get("lon"))
        item["rating"] = to_float(item.get("rating"))
        item["review_count"] = to_float(item.get("review_count"))
        item["price"] = parse_price(item.get("price_level"))
        cleaned.append(item)
    return cleaned


def validate_rows(rows: list[dict[str, Any]], source_columns: set[str]) -> tuple[list[str], list[dict[str, Any]]]:
    errors: list[str] = []
    warnings: list[dict[str, Any]] = []
    missing_columns = sorted(REQUIRED_COLUMNS - source_columns)
    if missing_columns:
        errors.append(f"Missing required columns: {', '.join(missing_columns)}")
        return errors, warnings

    seen_urls: dict[str, int] = {}
    for index, row in enumerate(rows, start=2):
        for field in ("name", "primary_cuisine", "url"):
            if not row.get(field):
                errors.append(f"Row {index}: {field} is required")
        lat, lon = row.get("lat"), row.get("lon")
        if lat is None or not 12.0 <= lat <= 15.0:
            errors.append(f"Row {index}: latitude is missing or outside the Bangkok region")
        if lon is None or not 99.0 <= lon <= 102.0:
            errors.append(f"Row {index}: longitude is missing or outside the Bangkok region")
        rating = row.get("rating")
        if rating is None or not 0 <= rating <= 5:
            errors.append(f"Row {index}: rating must be between 0 and 5")
        reviews = row.get("review_count")
        if reviews is None or reviews < 0 or not reviews.is_integer():
            errors.append(f"Row {index}: review_count must be a non-negative integer")
        if row.get("price") is None:
            errors.append(f"Row {index}: price_level could not be parsed")
        url = row.get("url")
        if url:
            if url in seen_urls:
                errors.append(f"Row {index}: duplicate url (first seen on row {seen_urls[url]})")
            else:
                seen_urls[url] = index

    for field in sorted(OPTIONAL_COLUMNS):
        count = sum(not row.get(field) for row in rows)
        if count:
            warnings.append({"code": "missing_optional", "field": field, "count": count})

    duplicate_names = sorted(name for name, count in Counter(row.get("name") for row in rows).items() if name and count > 1)
    if duplicate_names:
        warnings.append({"code": "duplicate_name", "count": len(duplicate_names), "values": duplicate_names})
    return errors, warnings


def scale(values: list[float]) -> list[float]:
    low, high = min(values), max(values)
    if high == low:
        return [0.0] * len(values)
    return [(value - low) / (high - low) for value in values]


def build_records(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ratings = [float(row["rating"]) for row in rows]
    global_mean = sum(ratings) / len(ratings)
    minimum_reviews = 50
    weighted = [
        (row["rating"] * row["review_count"] + global_mean * minimum_reviews)
        / (row["review_count"] + minimum_reviews)
        for row in rows
    ]
    rating_norm = scale(weighted)
    review_logs = [math.log(row["review_count"] + 1) for row in rows]
    review_norm = scale(review_logs)

    records: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        record = {
            "name": row["name"],
            "lat": row["lat"],
            "lon": row["lon"],
            "primary_cuisine": row["primary_cuisine"],
            "cuisine_subtype": row.get("cuisine_subtype"),
            "rating": row["rating"],
            "weighted_rating": weighted[index],
            "rating_norm": rating_norm[index],
            "review_count": int(row["review_count"]),
            "review_weight_norm": review_norm[index],
            "price_min": row["price"][0],
            "price_max": row["price"][1],
            "price_mid": row["price"][2],
            "address": row.get("address"),
            "url": row["url"],
            "image_url": row.get("image_url"),
        }
        records.append({column: record[column] for column in EXPORT_COLUMNS})
    return records


def relative_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path.resolve())


def build_dataset(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
    raw_rows = load_rows(path)
    source_columns = {str(column).strip() for row in raw_rows for column in row}
    rows = clean_rows(raw_rows)
    errors, warnings = validate_rows(rows, source_columns)
    if errors:
        return [], {"valid": False, "errors": errors, "warnings": warnings}, errors

    records = build_records(rows)
    report = {
        "schema_version": 2,
        "valid": True,
        "source": relative_path(path),
        "source_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "source_modified_date": source_modified_date(path),
        "restaurant_count": len(records),
        "primary_cuisine_count": len({record["primary_cuisine"] for record in records}),
        "primary_cuisines": dict(sorted(Counter(record["primary_cuisine"] for record in records).items())),
        "warnings": warnings,
        "errors": [],
    }
    return records, report, []


def encode_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2, allow_nan=False) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Canonical .xlsx or .csv source")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Generated restaurant JSON")
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT, help="Generated quality report")
    parser.add_argument("--check", action="store_true", help="Fail if generated files are stale; do not write")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        print(f"Data source not found: {args.input}", file=sys.stderr)
        return 2

    records, report, errors = build_dataset(args.input)
    if errors:
        print("Data validation failed:", file=sys.stderr)
        for error in errors:
            print(f"- {error}", file=sys.stderr)
        return 1

    output_text = encode_json(records)
    report_text = encode_json(report)
    if args.check:
        stale = []
        for path, expected in ((args.output, output_text), (args.report, report_text)):
            if not path.exists() or path.read_text(encoding="utf-8") != expected:
                stale.append(relative_path(path))
        if stale:
            print(f"Generated data is stale: {', '.join(stale)}", file=sys.stderr)
            print("Run: python3 scripts/build_data.py", file=sys.stderr)
            return 1
        print(f"Data is valid and current: {len(records)} restaurants")
        return 0

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(output_text, encoding="utf-8")
    args.report.write_text(report_text, encoding="utf-8")
    print(f"Built {relative_path(args.output)} with {len(records)} restaurants")
    print(f"Quality report: {relative_path(args.report)} ({len(report['warnings'])} warnings)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
